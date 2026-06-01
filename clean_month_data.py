"""
clean_month_data.py

Turn a raw Badge Access Log export into an R&D Employee Clock-In Times
workbook (same layout as finished.xlsx).

Usage:
    python clean_month_data.py "C:\\path\\to\\Badge Access Logs - 5-11 to 5-22.xlsx"
    python clean_month_data.py log1.xlsx log2.xlsx --output-dir C:\\Users\\You\\Downloads

Output file name (auto from data dates):
    R&D Employee Clock-In Times - MM-DD-YYYY to MM-DD-YYYY.xlsx

Sheets:
    - Daily Detail   : first badge scan per employee per day + signed MinutesLate
    - Averages       : Employee, Average Clock-In Time, Average Minutes Late,
                       Scheduled Start Time
    - Missing Days   : weekdays in range with no scan on record
"""

from __future__ import annotations

import argparse
from datetime import datetime, time
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

# ---------------------------------------------------------------------------
# R&D roster — scheduled start times (24-hour HH:MM:SS)
# ---------------------------------------------------------------------------
SCHEDULED_START_TIMES: dict[str, str] = {
    "Amenah Rashid": "08:00:00",
    "Bo Tian": "07:00:00",
    "Brian He": "08:00:00",
    "Di Yao": "09:00:00",
    "Jason Guo": "09:00:00",
    "Jeffry Irace": "07:00:00",
    "Jolie Yang": "09:00:00",
    "Junjie Chen": "08:00:00",
    "Junjie Wu": "09:00:00",
    "Li Yuan": "08:00:00",
    "Mingqing Xu": "08:00:00",
    "Nur Ugurlu": "08:00:00",
    "Qiaozheng Liu": "08:00:00",
    "Qiqi Liang": "09:00:00",
    "Reyhan Barindik": "08:00:00",
    "Ryan Gumbayan": "07:00:00",
    "Terry Liu": "09:00:00",
    "Tianyi Wang": "08:00:00",
    "Tom Tang": "09:00:00",
    "Xuefei Meng": "08:00:00",
    "Yanina Yang": "09:00:00",
}


def _to_time(value) -> time:
    """Convert Excel/time strings/datetimes into a datetime.time object."""
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    return pd.to_datetime(str(value)).time()


def _is_junk_header_row(row_values: list) -> bool:
    """True when row 1 is a placeholder like Column1, Column2, Column3."""
    cells = [
        str(value).strip().lower().replace(" ", "")
        for value in row_values
        if pd.notna(value)
    ]
    if not cells:
        return False
    return all(cell.startswith("column") for cell in cells[:3])


def read_badge_logs(paths: list[Path]) -> pd.DataFrame:
    """
    Read one or more raw badge exports and return combined scan rows.

    Expected columns after cleanup:
        Time — e.g. "May 11, 2026, 06:50:35"
        User — employee name (from "Actor User" in the export)
    """
    frames: list[pd.DataFrame] = []

    for path in paths:
        if not path.exists():
            raise FileNotFoundError(f"Badge log not found: {path}")

        preview = pd.read_excel(path, header=None, nrows=2)
        header_row = 0 if not _is_junk_header_row(preview.iloc[0].tolist()) else 1
        logs = pd.read_excel(path, header=header_row)

        logs = logs.rename(columns={"Actor User": "User", "Actor user": "User"})

        required = {"Time", "User"}
        missing = required - set(logs.columns)
        if missing:
            raise ValueError(
                f"{path.name} is missing columns {sorted(missing)}. "
                f"Found: {list(logs.columns)}"
            )

        logs["Time"] = pd.to_datetime(logs["Time"], format="%b %d, %Y, %H:%M:%S")
        logs = logs.dropna(subset=["Time", "User"])
        logs["User"] = logs["User"].astype(str).str.strip()
        frames.append(logs[["Time", "User"]])

    combined = pd.concat(frames, ignore_index=True)
    combined = combined.drop_duplicates(subset=["Time", "User"])
    return combined.sort_values("Time").reset_index(drop=True)


def load_schedules(schedule_reference: Path | None) -> dict[str, time]:
    """
    Load scheduled start times.

    Uses SCHEDULED_START_TIMES by default. If schedule_reference points to an
    existing Averages sheet, those scheduled times override the defaults.
    """
    schedules: dict[str, time] = {
        name: _to_time(value) for name, value in SCHEDULED_START_TIMES.items()
    }

    if schedule_reference and schedule_reference.exists():
        preview = pd.read_excel(
            schedule_reference, sheet_name="Averages", header=None, nrows=2
        )
        header_row = 0 if not _is_junk_header_row(preview.iloc[0].tolist()) else 1
        averages = pd.read_excel(
            schedule_reference, sheet_name="Averages", header=header_row
        )
        for _, row in averages.iterrows():
            employee = str(row["Employee"]).strip()
            if employee.lower() in {"employee", "nan"}:
                continue
            schedules[employee] = _to_time(row["Scheduled Start Time"])

    return schedules


def first_clock_in_per_day(logs: pd.DataFrame, employees: set[str]) -> pd.DataFrame:
    """Earliest badge scan per R&D employee per calendar day."""
    logs = logs[logs["User"].isin(employees)].copy()
    logs["Date"] = logs["Time"].dt.normalize()

    daily = (
        logs.groupby(["User", "Date"], as_index=False)
        .first()[["User", "Date", "Time"]]
        .sort_values(["User", "Date"])
        .reset_index(drop=True)
    )
    daily["ClockIn"] = daily["Time"].dt.strftime("%H:%M:%S")
    return daily


def _clock_in_to_seconds(clock_in) -> int:
    """Seconds since midnight for a ClockIn value."""
    if isinstance(clock_in, time):
        parsed = clock_in
    elif isinstance(clock_in, datetime):
        parsed = clock_in.time()
    else:
        parsed = pd.to_datetime(str(clock_in), format="%H:%M:%S").time()
    return parsed.hour * 3600 + parsed.minute * 60 + parsed.second


def _seconds_to_hms(total_seconds: int) -> str:
    """Format seconds since midnight as HH:MM:SS."""
    total_seconds = int(total_seconds) % (24 * 3600)
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    secs = total_seconds % 60
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"


def compute_minutes_late(
    clock_in: time, work_date: pd.Timestamp, scheduled_start: time
) -> int:
    """
    Signed minutes vs scheduled start (Daily Detail 'MinutesLate').

    Negative = early, positive = late, zero = on time.
    """
    clock_dt = pd.Timestamp.combine(work_date.date(), clock_in)
    sched_dt = pd.Timestamp.combine(work_date.date(), scheduled_start)
    diff_minutes = (clock_dt - sched_dt).total_seconds() / 60
    return int(round(diff_minutes))


def build_daily_detail(daily: pd.DataFrame, schedules: dict[str, time]) -> pd.DataFrame:
    """Add signed MinutesLate to each daily clock-in row."""
    detail = daily[["User", "Date", "Time", "ClockIn"]].copy()
    detail["MinutesLate"] = detail.apply(
        lambda row: compute_minutes_late(
            row["Time"].time(), row["Date"], schedules[row["User"]]
        ),
        axis=1,
    )
    return detail


def build_averages(
    daily_detail: pd.DataFrame, schedules: dict[str, time]
) -> pd.DataFrame:
    """Summarize each employee from Daily Detail."""
    rows = []
    for employee, group in daily_detail.groupby("User"):
        seconds = [_clock_in_to_seconds(value) for value in group["ClockIn"]]
        mean_seconds = sum(seconds) / len(seconds)
        rows.append(
            {
                "Employee": employee,
                "Average Clock-In Time": _seconds_to_hms(mean_seconds),
                "Average Minutes Late": int(round(group["MinutesLate"].mean())),
                "Scheduled Start Time": schedules[employee],
            }
        )
    return pd.DataFrame(rows).sort_values("Employee").reset_index(drop=True)


def build_missing_days(
    daily: pd.DataFrame,
    schedules: dict[str, time],
    start_date: pd.Timestamp,
    end_date: pd.Timestamp,
) -> pd.DataFrame:
    """Weekdays in the report window with no badge scan."""
    business_days = pd.bdate_range(start_date, end_date)
    rows = []
    for employee in sorted(schedules):
        present_days = set(daily.loc[daily["User"] == employee, "Date"])
        for workday in business_days:
            if workday not in present_days:
                rows.append({"Employee": employee, "Missing Workday": workday})
    return pd.DataFrame(rows)


def output_filename(start_date: pd.Timestamp, end_date: pd.Timestamp) -> str:
    """R&D Employee Clock-In Times - MM-DD-YYYY to MM-DD-YYYY.xlsx"""
    start_label = start_date.strftime("%m-%d-%Y")
    end_label = end_date.strftime("%m-%d-%Y")
    return f"R&D Employee Clock-In Times - {start_label} to {end_label}.xlsx"


def _as_excel_datetime(value) -> datetime:
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, datetime):
        return value
    return pd.to_datetime(value).to_pydatetime()


def _average_minutes_late_formula(detail_last_row: int, row: int) -> str:
    """Excel formula: average MinutesLate from Daily Detail for this employee."""
    return (
        f"=ROUND(AVERAGEIFS('Daily Detail'!$E$2:$E${detail_last_row},"
        f"'Daily Detail'!$A$2:$A${detail_last_row},A{row}),0)"
    )


def write_report(
    daily_detail: pd.DataFrame,
    averages: pd.DataFrame,
    missing_days: pd.DataFrame,
    output_path: Path,
) -> None:
    """Write the three-sheet workbook (finished.xlsx layout)."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    detail_sheet = workbook.create_sheet("Daily Detail")
    detail_sheet.append(["User", "Date", "Time", "ClockIn", "MinutesLate"])
    for _, row in daily_detail.iterrows():
        detail_sheet.append(
            [
                row["User"],
                _as_excel_datetime(row["Date"]),
                _as_excel_datetime(row["Time"]),
                row["ClockIn"],
                int(row["MinutesLate"]),
            ]
        )

    detail_last_row = len(daily_detail) + 1

    averages_sheet = workbook.create_sheet("Averages")
    averages_sheet.append(
        [
            "Employee",
            "Average Clock-In Time",
            "Average Minutes Late",
            "Scheduled Start Time",
        ]
    )
    for row_idx, (_, row) in enumerate(averages.iterrows(), start=2):
        averages_sheet.append(
            [
                row["Employee"],
                row["Average Clock-In Time"],
                _average_minutes_late_formula(detail_last_row, row_idx),
                row["Scheduled Start Time"],
            ]
        )

    missing_sheet = workbook.create_sheet("Missing Days")
    missing_sheet.append(["Employee", "Missing Workday"])
    for _, row in missing_days.iterrows():
        missing_sheet.append(
            [row["Employee"], _as_excel_datetime(row["Missing Workday"])]
        )

    workbook.save(output_path)
    workbook.close()


def process_month(
    input_paths: list[Path],
    output_dir: Path | None = None,
    schedule_reference: Path | None = None,
) -> Path:
    """
    Build the monthly clock-in report from raw badge log file(s).

    Returns the path to the written Excel file.
    """
    logs = read_badge_logs(input_paths)
    schedules = load_schedules(schedule_reference)
    employees = set(schedules.keys())

    daily = first_clock_in_per_day(logs, employees)
    if daily.empty:
        raise SystemExit("No R&D badge records found in the input file(s).")

    start_date = daily["Date"].min()
    end_date = daily["Date"].max()

    daily_detail = build_daily_detail(daily, schedules)
    averages = build_averages(daily_detail, schedules)
    missing_days = build_missing_days(daily, schedules, start_date, end_date)

    destination = output_dir if output_dir else input_paths[0].parent
    destination.mkdir(parents=True, exist_ok=True)
    output_path = destination / output_filename(start_date, end_date)

    write_report(daily_detail, averages, missing_days, output_path)
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert raw Badge Access Logs into an R&D clock-in report."
    )
    parser.add_argument(
        "inputs",
        nargs="+",
        type=Path,
        help="One or more raw badge access log .xlsx files for the month/period",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help="Folder for the output file (default: same folder as the first input)",
    )
    parser.add_argument(
        "--schedule-reference",
        type=Path,
        default=None,
        help="Optional existing report whose Averages sheet supplies scheduled start times",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_paths = [path.resolve() for path in args.inputs]

    print("Input badge log(s):")
    for path in input_paths:
        print(f"  - {path}")

    output_path = process_month(
        input_paths,
        output_dir=args.output_dir.resolve() if args.output_dir else None,
        schedule_reference=(
            args.schedule_reference.resolve() if args.schedule_reference else None
        ),
    )

    logs = read_badge_logs(input_paths)
    print(f"Raw data range  : {logs['Time'].min().date()} to {logs['Time'].max().date()}")
    print(f"Report saved to : {output_path}")


if __name__ == "__main__":
    main()
